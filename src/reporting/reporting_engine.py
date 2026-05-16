"""
Enterprise operational reporting — Excel workbooks from gold marts and audits.

Uses ``xlsxwriter`` for primary workbook construction and ``openpyxl`` for
exception / audit workbooks with conditional formatting and traceability views.
"""

from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
import xlsxwriter
from xlsxwriter.utility import xl_col_to_name
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from utils.logger import get_logger

from .master_workbook_builder import build_master_executive_workbook

_LOG = get_logger("reporting")


def _resolve_path(project_root: Path, configured: str | Path) -> Path:
    path = Path(configured)
    return path if path.is_absolute() else (project_root / path)


def _read_csv(path: Path) -> pd.DataFrame:
    if not path.is_file():
        return pd.DataFrame()
    try:
        return pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8", on_bad_lines="skip")
    except (pd.errors.EmptyDataError, pd.errors.ParserError) as exc:
        _LOG.warning("Unreadable or empty CSV %s: %s", path, exc)
        return pd.DataFrame()


def _pct(num: float, den: float) -> float:
    if den <= 0:
        return 0.0
    return round(100.0 * float(num) / float(den), 2)


@dataclass
class ReportingEngine:
    """
    Build operational Excel reports under ``outputs/excel_reports/``.

    Args:
        project_root: Repository root.
        config: Parsed ``config.yaml``.
        processing_batch_id: Current validation / gold batch id.
    """

    project_root: Path
    config: Mapping[str, Any]
    processing_batch_id: str | None = None
    report_paths: dict[str, Path] = field(default_factory=dict)

    def __post_init__(self) -> None:
        rcfg = self.config.get("reporting") or {}
        out_base = _resolve_path(self.project_root, self.config.get("output_path", "outputs"))
        self._excel_dir = out_base / rcfg.get("excel_output_dir", "excel_reports")
        self._row_cap = int(rcfg.get("audit_detail_row_cap", 50000))
        self._executive = self._excel_dir / rcfg.get("executive_workbook", "Executive_Operations_Summary.xlsx")
        self._recon_wb = self._excel_dir / rcfg.get("reconciliation_workbook", "Reconciliation_Report.xlsx")
        self._validation_wb = self._excel_dir / rcfg.get("validation_workbook", "Validation_Exception_Report.xlsx")
        self._dup_wb = self._excel_dir / rcfg.get("duplicate_workbook", "Duplicate_Intelligence_Report.xlsx")
        self._audit_wb = self._excel_dir / rcfg.get("audit_workbook", "Audit_Traceability_Report.xlsx")
        self._master_wb = self._excel_dir / rcfg.get(
            "master_executive_workbook",
            "Enterprise_Banking_Operations_Master_Report.xlsx",
        )

        gl = self.config.get("gold_layer") or {}
        gold_dir = _resolve_path(self.project_root, self.config.get("gold_path", "data/gold"))
        self._txn_master = gold_dir / gl.get("transaction_master_filename", "gold_transaction_master.csv")
        self._dup_mart = gold_dir / gl.get("duplicate_summary_filename", "gold_duplicate_summary.csv")
        self._val_mart = gold_dir / gl.get("validation_summary_filename", "gold_validation_summary.csv")
        self._rec_mart = gold_dir / gl.get("reconciliation_summary_filename", "gold_reconciliation_summary.csv")
        self._fx_mart = gold_dir / gl.get("fx_variance_summary_filename", "gold_fx_variance_summary.csv")

        val_cfg = self.config.get("validation") or {}
        self._validation_audit = _resolve_path(
            self.project_root, self.config.get("audit_path", "data/audit")
        ) / val_cfg.get("validation_audit_log_filename", "validation_audit_log.csv")

        rec_cfg = self.config.get("reconciliation") or {}
        self._rec_audit = _resolve_path(
            self.project_root, self.config.get("audit_path", "data/audit")
        ) / rec_cfg.get("reconciliation_audit_log_filename", "reconciliation_audit_log.csv")

        dup_cfg = self.config.get("duplicate_intelligence") or {}
        self._dup_detail = out_base / dup_cfg.get(
            "duplicate_detail_path", "profiling_reports/duplicate_intelligence_detailed.csv"
        )

    def run(
        self,
        mart_paths: Mapping[str, Path] | None = None,
        *,
        pipeline_duration_seconds: float = 0.0,
    ) -> dict[str, Path]:
        self._excel_dir.mkdir(parents=True, exist_ok=True)
        txn_path = (mart_paths or {}).get("transaction_master") or self._txn_master
        dup_path = (mart_paths or {}).get("duplicate_summary") or self._dup_mart
        val_path = (mart_paths or {}).get("validation_summary") or self._val_mart
        rec_path = (mart_paths or {}).get("reconciliation_summary") or self._rec_mart
        fx_path = (mart_paths or {}).get("fx_variance_summary") or self._fx_mart

        txn = _read_csv(Path(txn_path))
        dup_m = _read_csv(Path(dup_path))
        val_m = _read_csv(Path(val_path))
        rec_m = _read_csv(Path(rec_path))
        fx_m = _read_csv(Path(fx_path))

        metrics = self._compute_metrics(txn, val_m, rec_m, dup_m, pipeline_duration_seconds)

        self._build_executive_xlsx(metrics, txn, val_m, rec_m, dup_m, fx_m)
        self._build_reconciliation_xlsx(txn, _read_csv(self._rec_audit), rec_m)
        self._build_duplicate_xlsx(dup_m, _read_csv(self._dup_detail))
        self._build_validation_openpyxl(_read_csv(self._validation_audit))
        self._build_audit_openpyxl(_read_csv(self._rec_audit), _read_csv(self._validation_audit))

        build_master_executive_workbook(
            project_root=self.project_root,
            config=self.config,
            processing_batch_id=self.processing_batch_id,
            pipeline_duration_seconds=pipeline_duration_seconds,
            output_path=self._master_wb,
        )

        self.report_paths = {
            "executive": self._executive,
            "reconciliation": self._recon_wb,
            "validation": self._validation_wb,
            "duplicate": self._dup_wb,
            "audit": self._audit_wb,
            "master_executive": self._master_wb,
        }
        _LOG.info("Reporting workbooks written to %s", self._excel_dir)
        return self.report_paths

    def _compute_metrics(
        self,
        txn: pd.DataFrame,
        val_m: pd.DataFrame,
        rec_m: pd.DataFrame,
        dup_m: pd.DataFrame,
        pipeline_duration_seconds: float,
    ) -> dict[str, Any]:
        batch = self.processing_batch_id or ""
        n_txn = len(txn)
        tot_proc = int(float(val_m.iloc[0]["total_rows_processed"])) if not val_m.empty else n_txn
        tot_val = int(float(val_m.iloc[0]["total_valid"])) if not val_m.empty else 0
        tot_rej = int(float(val_m.iloc[0]["total_rejected"])) if not val_m.empty else 0
        tot_warn = int(float(val_m.iloc[0]["total_warning"])) if not val_m.empty else 0
        total_matched = int(float(rec_m.iloc[0]["total_matched"])) if not rec_m.empty else 0
        total_rows_rec = int(float(rec_m.iloc[0]["total_rows"])) if not rec_m.empty else n_txn
        dup_pairs = int(dup_m["pair_count"].fillna(0).astype(float).sum()) if not dup_m.empty else 0
        unresolved = 0
        if not rec_m.empty:
            unresolved = int(float(rec_m.iloc[0].get("MISSING_SETTLEMENT", 0) or 0)) + int(
                float(rec_m.iloc[0].get("MULTIPLE_MATCHES", 0) or 0)
            )
        return {
            "processing_batch_id": batch,
            "total_transactions_processed": tot_proc or n_txn,
            "gold_master_row_count": n_txn,
            "validation_success_rate_pct": _pct(tot_val, tot_proc) if tot_proc else 0.0,
            "reconciliation_success_rate_pct": _pct(total_matched, total_rows_rec) if total_rows_rec else 0.0,
            "duplicate_detection_pairs": dup_pairs,
            "unresolved_reconciliation_rows": unresolved,
            "rejected_record_count": tot_rej,
            "warning_record_count": tot_warn,
            "pipeline_duration_seconds": round(float(pipeline_duration_seconds), 4),
        }

    def _build_executive_xlsx(
        self,
        metrics: dict[str, Any],
        txn: pd.DataFrame,
        val_m: pd.DataFrame,
        rec_m: pd.DataFrame,
        dup_m: pd.DataFrame,
        fx_m: pd.DataFrame,
    ) -> None:
        wb = xlsxwriter.Workbook(str(self._executive))
        fmt_title = wb.add_format({"bold": True, "font_size": 14, "bg_color": "#1F4E79", "font_color": "white"})
        fmt_hdr = wb.add_format({"bold": True, "bg_color": "#D9E1F2", "border": 1})
        fmt_cell = wb.add_format({"border": 1})
        fmt_num = wb.add_format({"border": 1, "num_format": "0.00"})
        fmt_pct = wb.add_format({"border": 1, "num_format": "0.00%"})

        kpi = wb.add_worksheet("Executive_KPIs")
        kpi.set_column("A:A", 38)
        kpi.set_column("B:B", 22)
        kpi.merge_range("A1:B1", "Executive Operations Summary", fmt_title)
        kpi.freeze_panes(3, 0)
        row = 2
        pairs = [
            ("Processing batch", metrics.get("processing_batch_id", "")),
            ("Total transactions processed", metrics.get("total_transactions_processed", 0)),
            ("Gold master row count", metrics.get("gold_master_row_count", 0)),
            ("Validation success rate (%)", metrics.get("validation_success_rate_pct", 0) / 100.0),
            ("Reconciliation success rate (%)", metrics.get("reconciliation_success_rate_pct", 0) / 100.0),
            ("Duplicate detection (pair rows in mart)", metrics.get("duplicate_detection_pairs", 0)),
            ("Unresolved reconciliation rows", metrics.get("unresolved_reconciliation_rows", 0)),
            ("Rejected records (validation)", metrics.get("rejected_record_count", 0)),
            ("Warning records (validation)", metrics.get("warning_record_count", 0)),
            ("Pipeline duration (seconds)", metrics.get("pipeline_duration_seconds", 0)),
        ]
        for lab, val in pairs:
            kpi.write(row, 0, lab, fmt_hdr)
            if "rate" in lab.lower():
                kpi.write_number(row, 1, float(val), fmt_pct)
            elif isinstance(val, (int, float)) and "batch" not in lab.lower():
                kpi.write_number(row, 1, float(val), fmt_num)
            else:
                kpi.write(row, 1, str(val), fmt_cell)
            row += 1
        kpi.conditional_format(
            "B4:B5",
            {"type": "3_color_scale", "min_color": "#F8696B", "mid_color": "#FFEB84", "max_color": "#63BE7B"},
        )

        summ = wb.add_worksheet("Mart_Summary")
        summ.write_row(0, 0, ["Mart", "Path", "Row_count"], fmt_hdr)
        summ.freeze_panes(1, 0)
        r = 1
        for name, df, p in (
            ("gold_transaction_master", txn, self._txn_master),
            ("gold_validation_summary", val_m, self._val_mart),
            ("gold_reconciliation_summary", rec_m, self._rec_mart),
            ("gold_duplicate_summary", dup_m, self._dup_mart),
            ("gold_fx_variance_summary", fx_m, self._fx_mart),
        ):
            summ.write_row(r, 0, [name, str(p), len(df)], fmt_cell)
            r += 1
        summ.autofilter(0, 0, max(0, r - 1), 2)
        summ.set_column(0, 0, 28)
        summ.set_column(1, 1, 72)
        summ.set_column(2, 2, 12)
        wb.close()

    def _build_reconciliation_xlsx(self, txn: pd.DataFrame, rec_aud: pd.DataFrame, rec_m: pd.DataFrame) -> None:
        wb = xlsxwriter.Workbook(str(self._recon_wb))
        fmt_hdr = wb.add_format({"bold": True, "bg_color": "#D9E1F2", "border": 1})
        fmt_cell = wb.add_format({"border": 1})

        ws0 = wb.add_worksheet("Summary")
        ws0.freeze_panes(1, 0)
        if not rec_m.empty:
            cols = list(rec_m.columns)
            ws0.write_row(0, 0, cols, fmt_hdr)
            ws0.write_row(1, 0, [rec_m.iloc[0].get(c, "") for c in cols], fmt_cell)
            last_r = len(rec_m)
            ws0.autofilter(0, 0, last_r, len(cols) - 1)
        else:
            ws0.write(0, 0, "No reconciliation summary available.", fmt_cell)

        ws1 = wb.add_worksheet("Reconciliation_Detail")
        ws1.freeze_panes(1, 0)
        if not rec_aud.empty:
            sub = rec_aud.head(self._row_cap)
            cols = list(sub.columns)
            ws1.write_row(0, 0, cols, fmt_hdr)
            for i, row in enumerate(sub.itertuples(index=False), start=1):
                ws1.write_row(i, 0, list(row), fmt_cell)
            ws1.autofilter(0, 0, len(sub), len(cols) - 1)
            for j, _ in enumerate(cols):
                ws1.set_column(j, j, 14)
            # confidence column conditional format if present
            if "confidence_score" in cols:
                cidx = cols.index("confidence_score")
                cl = xl_col_to_name(cidx)
                rng = f"{cl}2:{cl}{len(sub) + 1}"
                ws1.conditional_format(
                    rng,
                    {"type": "3_color_scale", "min_color": "#F8696B", "mid_color": "#FFEB84", "max_color": "#63BE7B"},
                )
        else:
            ws1.write(0, 0, "No reconciliation audit rows.", fmt_cell)

        ws2 = wb.add_worksheet("Transaction_Master_Sample")
        ws2.freeze_panes(1, 0)
        if not txn.empty:
            sub = txn.head(min(5000, len(txn)))
            cols = list(sub.columns)
            ws2.write_row(0, 0, cols, fmt_hdr)
            for i, row in enumerate(sub.itertuples(index=False), start=1):
                ws2.write_row(i, 0, list(row), fmt_cell)
            ws2.autofilter(0, 0, len(sub), len(cols) - 1)
        wb.close()

    def _build_duplicate_xlsx(self, dup_m: pd.DataFrame, dup_detail: pd.DataFrame) -> None:
        wb = xlsxwriter.Workbook(str(self._dup_wb))
        fmt_hdr = wb.add_format({"bold": True, "bg_color": "#D9E1F2", "border": 1})
        fmt_cell = wb.add_format({"border": 1})

        s0 = wb.add_worksheet("Duplicate_Summary")
        s0.freeze_panes(1, 0)
        if not dup_m.empty:
            cols = list(dup_m.columns)
            s0.write_row(0, 0, cols, fmt_hdr)
            for i, (_, row) in enumerate(dup_m.iterrows(), start=1):
                s0.write_row(i, 0, [row.get(c, "") for c in cols], fmt_cell)
            s0.autofilter(0, 0, len(dup_m), len(cols) - 1)
        s1 = wb.add_worksheet("Duplicate_Pairs")
        s1.freeze_panes(1, 0)
        if not dup_detail.empty:
            sub = dup_detail.head(self._row_cap)
            cols = list(sub.columns)
            s1.write_row(0, 0, cols, fmt_hdr)
            for i, row in enumerate(sub.itertuples(index=False), start=1):
                s1.write_row(i, 0, list(row), fmt_cell)
            s1.autofilter(0, 0, len(sub), len(cols) - 1)
            if "duplicate_confidence" in cols:
                cidx = cols.index("duplicate_confidence")
                cl = xl_col_to_name(cidx)
                rng = f"{cl}2:{cl}{len(sub) + 1}"
                s1.conditional_format(
                    rng,
                    {"type": "3_color_scale", "min_color": "#F8696B", "mid_color": "#FFEB84", "max_color": "#63BE7B"},
                )
        wb.close()

    def _build_validation_openpyxl(self, aud: pd.DataFrame) -> None:
        wb = Workbook()
        sum_ws = wb.active
        sum_ws.title = "Exception_Summary"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        sum_ws.append(["Metric", "Value"])
        sub_full = aud.head(self._row_cap)
        sum_ws.append(["Audit rows (capped sample)", len(sub_full)])
        if not sub_full.empty and "rule_name" in sub_full.columns:
            vc = sub_full["rule_name"].value_counts().head(25)
            sum_ws.append([])
            sum_ws.append(["Top rules", "Count"])
            for k, v in vc.items():
                sum_ws.append([str(k), int(v)])
        sum_ws.freeze_panes = "A2"
        sum_ws.column_dimensions["A"].width = 36
        sum_ws.column_dimensions["B"].width = 14

        ws = wb.create_sheet("Validation_Audit")
        if aud.empty:
            ws.append(["message"])
            ws.append(["No validation audit rows available."])
            wb.save(self._validation_wb)
            return

        sub = sub_full
        cols = list(sub.columns)
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for _, row in sub.iterrows():
            ws.append([row.get(col, "") for col in cols])

        last_row = len(sub) + 1
        last_col = len(cols)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

        for c in range(1, last_col + 1):
            try:
                col_vals = sub.iloc[:, c - 1].astype(str).head(200)
                mw = int(min(28, 10 + col_vals.map(len).max()))
            except ValueError:
                mw = 14
            ws.column_dimensions[get_column_letter(c)].width = mw

        sev_col = None
        for i, name in enumerate(cols, start=1):
            if str(name).lower() == "severity":
                sev_col = get_column_letter(i)
                break
        if sev_col:
            red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            amber = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            rng = f"{sev_col}2:{sev_col}{last_row}"
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f'{sev_col}2="S5"'], fill=red),
            )
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f'{sev_col}2="S4"'], fill=amber),
            )

        wb.save(self._validation_wb)

    def _build_audit_openpyxl(self, rec_aud: pd.DataFrame, val_aud: pd.DataFrame) -> None:
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="375623")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        sum_ws = wb.active
        sum_ws.title = "Traceability_Index"
        sum_ws.append(["Artifact", "Row count (capped)", "Traceability notes"])
        sum_ws.append(
            [
                "Reconciliation audit",
                str(min(len(rec_aud), self._row_cap)),
                "pipeline_row_id links to gold_transaction_master; matching_logic_used explains match.",
            ]
        )
        sum_ws.append(
            [
                "Validation audit",
                str(min(len(val_aud), self._row_cap)),
                "rule_name and severity tie to validation engine rules; source_file + row_identifier.",
            ]
        )
        sum_ws.freeze_panes = "A2"
        sum_ws.column_dimensions["A"].width = 28
        sum_ws.column_dimensions["B"].width = 22
        sum_ws.column_dimensions["C"].width = 72

        def write_sheet(name: str, df: pd.DataFrame) -> None:
            ws = wb.create_sheet(name)
            if df.empty:
                ws.append(["message"])
                ws.append([f"No rows in {name}."])
                return
            sub = df.head(self._row_cap)
            cols = list(sub.columns)
            ws.append(cols)
            for c in range(1, len(cols) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            for _, row in sub.iterrows():
                ws.append([row.get(col, "") for col in cols])
            last_row = len(sub) + 1
            last_col = len(cols)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

        write_sheet("Reconciliation_Audit", rec_aud)
        write_sheet("Validation_Audit", val_aud)

        wb.save(self._audit_wb)
